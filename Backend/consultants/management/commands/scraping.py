import re
import requests
from bs4 import BeautifulSoup
from pathlib import Path
import html
import ftfy
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.core.management.base import BaseCommand
from urllib.parse import urljoin
from datetime import datetime
import time
import socket
import logging
import traceback
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, 
    TimeoutException,
    ElementClickInterceptedException,
    WebDriverException,
    StaleElementReferenceException
)
from webdriver_manager.chrome import ChromeDriverManager
import csv

# Configuration du logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class Command(BaseCommand):
    help = 'Scrape Beta-Conseils, SNIM, RIMTIC, SOMELEC et Marchés Publics Mauritanie pour les appels d\'offres'

    def __init__(self):
        super().__init__()
        self.scraped_count = 0
        self.target_count = 300
        self.max_retries = 3
        self.delay_between_requests = 2
        self.session_restart_threshold = 25

    def add_arguments(self, parser):
        parser.add_argument(
            '--site',
            type=str,
            help='Scraper un site spécifique (are, beta, snim, rimtic, somelec, marchespublics, all)',
            default='all'
        )

    def handle(self, *args, **options):
        site = options['site'].lower() if options.get('site') else 'all'
        
        self.stdout.write("🚀 Démarrage du scraping...")
        self.stdout.write(f"Site(s) ciblé(s): {site}")
        self.stdout.write("Les résultats seront sauvegardés dans des fichiers séparés\n")
        if site == 'all' or site == 'rimtic':
            self.scrape_rimtic()
        if site == 'all' or site == 'somelec':
            self.scrape_somelec()
        if site == 'all' or site == 'marchespublics':
            self.scrape_marchespublics_mauritanie()
        if site == 'all' or site == 'beta':
            self.scrape_beta_conseils()
        if site == 'all' or site == 'snim':
            self.scrape_snim()
        
        
        
        self.stdout.write(self.style.SUCCESS("\n✅ Scraping terminé pour tous les sites sélectionnés!"))

    def check_internet_connection(self):
        try:
            socket.create_connection(("8.8.8.8", 53), timeout=5)
            return True
        except OSError:
            return False

    def clean_text(self, text):
        """Nettoyer le texte des espaces excessifs et caractères spéciaux"""
        if not text:
            return ""
        text = str(text)
        return " ".join(text.split()).strip()

    def clean_special_characters(self, text):
        if text is None:
            return None
        text = str(text)
        text = html.escape(text)
        text = ftfy.fix_text(text)
        return text

    def setup_chrome_driver(self):
        """Configuration du driver Chrome avec webdriver-manager"""
        try:
            options = Options()
            options.add_argument("--start-maximized")
            options.add_argument("--disable-infobars")
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-notifications")
            options.add_argument("--disable-popup-blocking")
            options.add_argument("--headless")
            options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            driver.implicitly_wait(10)
            
            logger.info("✅ Driver Chrome configuré avec succès")
            return driver
        except Exception as e:
            logger.error(f"❌ Erreur configuration driver: {e}")
            raise

    def cleanup_driver(self, driver):
        """Nettoyer proprement le driver"""
        if driver:
            try:
                driver.quit()
                logger.info("🔒 Driver fermé proprement")
            except Exception as e:
                logger.warning(f"⚠️ Erreur fermeture driver: {e}")


    

    def load_all_offers(self, driver):
        """Fonction pour charger toutes les offres en cliquant sur 'Voir plus'"""
        last_count = 0
        same_count = 0
        max_attempts = 30
        
        for attempt in range(1, max_attempts + 1):
            current_offers = driver.find_elements(By.CSS_SELECTOR, "h2.text-sm.mb-2.pt-6.text-green-950")
            current_count = len(current_offers)
            
            self.stdout.write(f"📊 Offres chargées: {current_count} (tentative {attempt}/{max_attempts})")
            
            if current_count == last_count:
                same_count += 1
                if same_count >= 5:
                    self.stdout.write("🔴 Blocage détecté, tentative de déblocage...")
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(3)
                    driver.execute_script("window.scrollTo(0, 0);")
                    time.sleep(2)
                    same_count = 0
            else:
                same_count = 0
                last_count = current_count
            
            try:
                buttons = WebDriverWait(driver, 15).until(
                    EC.presence_of_all_elements_located((By.XPATH, "//button[contains(., 'Voir plus')]"))
                )
                if buttons:
                    last_button = buttons[-1]
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", last_button)
                    time.sleep(1)
                    driver.execute_script("arguments[0].click();", last_button)
                    self.stdout.write(f"✅ Clic sur 'Voir plus' (tentative {attempt})")
                    time.sleep(3)
            except (NoSuchElementException, TimeoutException):
                self.stdout.write("ℹ️ Bouton 'Voir plus' non trouvé - fin du chargement")
                break
            except (ElementClickInterceptedException, StaleElementReferenceException):
                self.stdout.write("⚠️ Clic bloqué, tentative de récupération...")
                time.sleep(3)
                continue
            except Exception as e:
                self.stdout.write(f"❌ Erreur inattendue: {str(e)}")
                time.sleep(3)
                continue



    def scrape_marchespublics_mauritanie(self):
        """Scraping amélioré pour Marchés Publics Mauritanie"""
        self.stdout.write(self.style.SUCCESS("\n🔍 Démarrage du scraping Marchés Publics Mauritanie..."))
        self.stdout.write(f"🎯 Objectif minimum: {self.target_count} appels d'offres\n")
        
        OUTPUT_FILE = Path.cwd() / 'scraped_data' / 'marchespublics_mauritanie.xlsx'
        OUTPUT_FILE.parent.mkdir(exist_ok=True)

        driver = None
        workbook = None
        worksheet = None
        
        try:
            # Configuration du driver
            driver = self.setup_chrome_driver()
            
            # Configuration Excel
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Marchés Publics"
            
            # En-têtes avec les champs demandés
            headers = [
                "titre", 
                "date_de_publication", 
                "client", 
                "type_d_appel_d_offre",
                "date_limite", 
                "documents", 
                "lien_site"
            ]
            
            # Formatage des en-têtes
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Largeurs des colonnes
            column_widths = [50, 50, 50, 50, 50, 50, 50]
            for col, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = width

            # Accès au site
            url = "https://marchespublics.gov.mr/marchespublics"
            self.stdout.write(f"🌐 Accès à {url}...")
            driver.get(url)
            
            # Attendre le chargement initial
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.bg-white")))
            
            # Gestion des cookies
            try:
                cookie_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Accepter') or contains(., 'OK')]")))
                cookie_btn.click()
                self.stdout.write("✅ Cookies acceptés")
                time.sleep(1)
            except Exception:
                self.stdout.write("⚠️ Pas de popup cookies trouvée")

            # Charger toutes les offres
            self.load_all_offers(driver)
            
            # Extraire les données
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            offers = soup.find_all('div', class_='bg-white')
            self.stdout.write(f"📊 Nombre total d'offres trouvées: {len(offers)}")
            
            for offer in offers:
                # The code is checking if the `scraped_count` variable is equal to the `target_count` variable in a Python class. If they are equal, the condition is satisfied.
                # The code snippet is checking if the `scraped_count` is equal to the `target_count`. If they are equal, the condition will be true.
                if self.scraped_count == self.target_count:
                    break
                    
                try:
                    # Titre et lien
                    title_element = offer.find('h2', class_=lambda x: x and 'text-sm' in str(x))
                    if not title_element:
                        continue
                        
                    title = self.clean_text(title_element.get_text())
                    link = offer.find('a', href=True)
                    link = link['href'] if link else "N/A"
                    if link != "N/A" and not link.startswith('http'):
                        link = f"https://marchespublics.gov.mr{link}"

                    # Extraire le texte complet du conteneur
                    container_text = offer.get_text()
                    clean_text = self.clean_text(container_text)

                    # Nettoyer le texte des éléments de navigation
                    navigation_keywords = [
                        'Accueil', 'Marchés Publics', 'Filtrer par', 'Se connecter',
                        'Publications', 'Recours', 'Lois', 'Régulation', 'العربية',
                        'Open menu', 'Fermer', 'Retour', 'Menu'
                    ]
                    for keyword in navigation_keywords:
                        clean_text = clean_text.replace(keyword, '')
                    clean_text = re.sub(r'\s+', ' ', clean_text).strip()

                    # Client (Autorité contractante)
                    client = "N/A"
                    client_patterns = [
                        r'Autorité contractante\s*:\s*([^\n]+)',
                        r'Maître d\'ouvrage\s*:\s*([^\n]+)',
                        r'Client\s*:\s*([^\n]+)',
                        r'Organisme\s*:\s*([^\n]+)',
                        r'(Ministère\s+[^\n]+)',
                        r'(Direction\s+[^\n]+)',
                        r'(Société\s+[^\n]+)',
                        r'Autorité\s*:\s*([^\n]+)'
                    ]
                    
                    for pattern in client_patterns:
                        match = re.search(pattern, clean_text, re.IGNORECASE)
                        if match:
                            client = self.clean_text(match.group(1))
                            client = re.sub(r'(Type|Types).*', '', client, flags=re.IGNORECASE).strip()
                            client = re.sub(r'Date.*', '', client).strip()
                            break

                    # Type d'appel d'offre
                    market_type = "N/A"
                    type_patterns = [
                        r'Type[s]? de[s]? marché[s]?\s*:\s*([^\n]+)',
                        r'Type[s]?\s*:\s*([^\n]+)',
                        r'Nature\s*:\s*([^\n]+)',
                        r'Catégorie\s*:\s*([^\n]+)'
                    ]
                    
                    for pattern in type_patterns:
                        match = re.search(pattern, clean_text, re.IGNORECASE)
                        if match:
                            market_type = self.clean_text(match.group(1))
                            market_type = re.sub(r'Date.*', '', market_type).strip()
                            market_type = re.sub(r'voir.*', '', market_type, flags=re.IGNORECASE).strip()
                            break

                    if market_type == "N/A":
                        type_mapping = {
                            'travaux': 'Travaux',
                            'fourniture': 'Fournitures',
                            'service': 'Services',
                            'étude': 'Études',
                            'prestation': 'Services',
                            'consultation': 'Consultation',
                            'recrutement': 'Recrutement',
                            'intellectuel': 'Prestations intellectuelles',
                            'matériel': 'Fournitures'
                        }
                        
                        for keyword, mtype in type_mapping.items():
                            if keyword in clean_text.lower():
                                market_type = mtype
                                break

                    # Dates
                    dates = re.findall(r'\d{1,2}/\d{1,2}/\d{2,4}', clean_text)
                    date_pub = dates[0] if dates else "N/A"
                    date_limite = dates[1] if len(dates) > 1 else "N/A"

                    # Documents
                    documents = "N/A"
                    doc_sections = offer.find_all(['p', 'div'], class_=lambda x: x and 'document' in str(x).lower())
                    if doc_sections:
                        doc_texts = [self.clean_text(s.get_text()) for s in doc_sections]
                        documents = " | ".join([d for d in doc_texts if d != "N/A" and len(d) > 10])
                    else:
                        paragraphs = offer.find_all('p')
                        meaningful_paragraphs = []
                        for p in paragraphs[:4]:
                            text = self.clean_text(p.get_text())
                            if text != "N/A" and len(text) > 20 and not any(kw in text.lower() for kw in ['date', 'type', 'autorité']):
                                meaningful_paragraphs.append(text)
                        
                        if meaningful_paragraphs:
                            documents = " | ".join(meaningful_paragraphs[:2])

                    # Ajouter les données à la feuille Excel
                    worksheet.append([
                        title,
                        date_pub,
                        client,
                        market_type,
                        date_limite,
                        documents if documents else "N/A",
                        link
                    ])

                    self.scraped_count += 1
                    self.stdout.write(f"✅ Offre {self.scraped_count}: {title[:50]}...")
                    self.stdout.write(f"   Client: {client}")
                    self.stdout.write(f"   Type: {market_type}")
                    self.stdout.write(f"   Publié: {date_pub}")
                    self.stdout.write(f"   Limite: {date_limite}")
                    
                    if self.scraped_count % 20 == 0:
                        self.stdout.write(f"📊 Progression: {self.scraped_count} offres traitées")
                    
                    # Sauvegarde périodique
                    if self.scraped_count % 10 == 0:
                        workbook.save(OUTPUT_FILE)
                        
                except Exception as e:
                    self.stdout.write(f"⚠️ Erreur lors du traitement d'une offre: {e}")
                    continue

            # Sauvegarde finale
            workbook.save(OUTPUT_FILE)
            self.stdout.write(self.style.SUCCESS(
                f"\n🎉 Scraping Marchés Publics terminé avec succès!\n"
                f"📊 Total offres enregistrées: {self.scraped_count}\n"
                f"📁 Fichier généré: {OUTPUT_FILE}"
            ))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Erreur globale Marchés Publics: {e}"))
            traceback.print_exc()
        finally:
            if workbook:
                try:
                    workbook.save(OUTPUT_FILE)
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"❌ Erreur sauvegarde finale: {e}"))
            self.cleanup_driver(driver)

    def load_all_offers(self, driver):
        """Fonction pour charger toutes les offres en cliquant sur 'Voir plus'"""
        last_count = 0
        same_count = 0
        max_attempts = 30
        
        for attempt in range(1, max_attempts + 1):
            current_offers = driver.find_elements(By.CSS_SELECTOR, "div.bg-white")
            current_count = len(current_offers)
            
            self.stdout.write(f"📊 Offres chargées: {current_count} (tentative {attempt}/{max_attempts})")
            
            if current_count == last_count:
                same_count += 1
                if same_count >= 5:
                    self.stdout.write("🔴 Blocage détecté, tentative de déblocage...")
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(3)
                    driver.execute_script("window.scrollTo(0, 0);")
                    time.sleep(2)
                    same_count = 0
            else:
                same_count = 0
                last_count = current_count
            
            try:
                buttons = WebDriverWait(driver, 15).until(
                    EC.presence_of_all_elements_located((By.XPATH, "//button[contains(., 'Voir plus')]"))
                )
                if buttons:
                    last_button = buttons[-1]
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", last_button)
                    time.sleep(1)
                    driver.execute_script("arguments[0].click();", last_button)
                    self.stdout.write(f"✅ Clic sur 'Voir plus' (tentative {attempt})")
                    time.sleep(3)
            except (NoSuchElementException, TimeoutException):
                self.stdout.write("ℹ️ Bouton 'Voir plus' non trouvé - fin du chargement")
                break
            except (ElementClickInterceptedException, StaleElementReferenceException):
                self.stdout.write("⚠️ Clic bloqué, tentative de récupération...")
                time.sleep(3)
                continue
            except Exception as e:
                self.stdout.write(f"❌ Erreur inattendue: {str(e)}")
                time.sleep(3)
                continue

    def scrape_beta_conseils(self):
        """Fonction de scraping pour Beta-Conseils"""
        self.stdout.write("\n🔍 Démarrage du scraping Beta-Conseils...")
        
        OUTPUT_FILE = Path.cwd() / 'scraped_data' / 'betaconseils.xlsx'
        OUTPUT_FILE.parent.mkdir(exist_ok=True)

        BASE_URL = "https://beta-conseils.com"
        START_URL = BASE_URL + "/appels-offres"
        HEADERS = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
        }
        DELAY_BETWEEN_REQUESTS = 1

        def get_offer_types(url):
            if not self.check_internet_connection():
                self.stdout.write("❌ Aucun accès Internet.")
                return []
            
            try:
                response = requests.get(url, headers=HEADERS, timeout=10)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, 'html.parser')
                offer_types_section = soup.find('div', class_='card-body')

                if not offer_types_section:
                    self.stdout.write("❌ Section des types d'offres introuvable.")
                    return []

                offer_types = []
                for item in offer_types_section.find_all('li', class_='d-inline-block'):
                    link = item.find('a', class_='text-white')
                    if not link:
                        continue
                    type_name = link.text.strip()
                    if type_name.lower().startswith("tous"):
                        continue
                    endpoint = link.get('href')
                    if endpoint:
                        offer_types.append((type_name, endpoint))
                return offer_types

            except requests.RequestException as e:
                self.stdout.write(f"❌ Erreur récupération des types : {e}")
                return []

        def scrape_offer_details(detail_url):
            try:
                response = requests.get(detail_url, headers=HEADERS, timeout=10)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, 'html.parser')
                
                pub_date = "N/A"
                short_info = soup.find('div', class_='job-short-info')
                if short_info:
                    date_label = short_info.find('h1', class_='text-dark h5', string='Date de publication')
                    if date_label:
                        date_item = date_label.find_next('li')
                        if date_item:
                            pub_date = date_item.get_text(strip=True)
                
                if pub_date == "N/A":
                    pub_date_tag = soup.find('p', class_='alert alert-info print')
                    if pub_date_tag:
                        pub_date = pub_date_tag.get_text(strip=True)
                        pub_date = pub_date.replace("Date de publication:", "").strip()
                
                if pub_date == "N/A":
                    all_dates = soup.find_all(text=lambda t: 'publication' in str(t).lower())
                    for date_text in all_dates:
                        if 'publication' in date_text.lower():
                            pub_date = date_text.strip()
                            break
                
                return self.clean_special_characters(pub_date)

            except requests.RequestException as e:
                self.stdout.write(f"⚠️ Erreur lors de l'accès à la page détaillée: {detail_url} - {e}")
                return "N/A"

        def scrape_offers(url, offer_type, sheet):
            try:
                response = requests.get(url, headers=HEADERS, timeout=10)
                response.raise_for_status()
                soup = BeautifulSoup(response.text, 'html.parser')
                offers = soup.find_all('div', class_='job-ad-item')

                if not offers:
                    self.stdout.write(f"⚠️ Aucune offre trouvée pour : {offer_type}")
                    return

                for offer in offers:
                    title_tag = offer.find('a', class_='job-title hover-blue')
                    client_tag = offer.find('a', class_='text-success d-inline-block mt-1')
                    deadline_tag = offer.find('a', class_='text-danger-2')
                    
                    title = self.clean_special_characters(title_tag.text.strip()) if title_tag else "N/A"
                    client = self.clean_special_characters(client_tag.text.strip()) if client_tag else "N/A"
                    deadline = " ".join(deadline_tag.stripped_strings) if deadline_tag else "N/A"
                    deadline = self.clean_special_characters(deadline)
                    
                    detail_url = BASE_URL + title_tag['href'] if title_tag and title_tag.has_attr('href') else "N/A"
                    
                    pub_date = "N/A"
                    if detail_url != "N/A":
                        pub_date = scrape_offer_details(detail_url)
                        time.sleep(DELAY_BETWEEN_REQUESTS)
                    
                    sheet.append([
                        title, pub_date, client, offer_type, deadline, "N/A", detail_url
                    ])

                    self.stdout.write(f"✅ {title[:50]}... | Client: {client} | Publié: {pub_date} | Clôture: {deadline}")

            except requests.RequestException as e:
                self.stdout.write(f"❌ Erreur scraping {offer_type} : {e}")

        if not self.check_internet_connection():
            self.stdout.write("❌ Pas de connexion Internet disponible.")
            return

        try:
            requests.get(BASE_URL, headers=HEADERS, timeout=10)
        except requests.RequestException as e:
            self.stdout.write(f"❌ Site inaccessible : {BASE_URL}\nDétails : {e}")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Appels d'offres"
        ws.append([
            "titre", "date_de_publication", "client", "type_d_appel_d_offre",
            "date_limite", "documents", "lien_site"
        ])

        offer_types = get_offer_types(START_URL)

        if not offer_types:
            self.stdout.write("❌ Aucun type d'offre récupéré.")
        else:
            for offer_type, endpoint in offer_types:
                url = BASE_URL + endpoint
                self.stdout.write(f"\n🔍 Scraping des offres pour : {offer_type}")
                scrape_offers(url, offer_type, ws)
                self.stdout.write("--------------------------------------------------")

        try:
            wb.save(OUTPUT_FILE)
            self.stdout.write(self.style.SUCCESS(f"\n✅ Résultat Beta-Conseils enregistré dans '{OUTPUT_FILE}'."))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Erreur sauvegarde Excel : {e}"))

    def scrape_snim(self):
        """Fonction de scraping pour SNIM"""
        self.stdout.write("\n🔍 Démarrage du scraping SNIM...")
        
        OUTPUT_FILE = Path.cwd() / 'scraped_data' / 'snim_offres.xlsx'
        OUTPUT_FILE.parent.mkdir(exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Appels d'offres SNIM"
        ws.append([
            "titre", "date_de_publication", "client", "type_d_appel_d_offre",
            "date_limite", "documents", "lien_site"
        ])

        BASE_URL = "https://www.snim.com"

        try:
            page_number = 0
            while True:
                url = f"{BASE_URL}/fr/ap-v?page={page_number}"
                self.stdout.write(f"Scraping page {page_number}...")

                response = requests.get(url, timeout=10)
                response.raise_for_status()

                soup = BeautifulSoup(response.text, 'html.parser')
                table = soup.find('table', class_='cols-3')

                if not table:
                    self.stdout.write("Fin de la pagination.")
                    break

                rows = table.find_all('tr', class_='apel')
                if len(rows) == 0:
                    self.stdout.write("Aucune donnée trouvée sur cette page.")
                    break

                for row in rows:
                    try:
                        cells = row.find_all('td')
                        if len(cells) >= 3:
                            titre_el = cells[0].find('a')
                            titre = self.clean_special_characters(titre_el.text.strip()) if titre_el else self.clean_special_characters(cells[0].text.strip())
                            lien_site = BASE_URL + titre_el['href'] if titre_el and titre_el.has_attr('href') else "N/A"

                            date_publication = self.clean_special_characters(cells[1].text.strip())
                            date_limite = self.clean_special_characters(cells[2].text.strip())
                            
                            client = "SNIM"
                            type_appel = "N/A"
                            documents = "N/A"

                            ws.append([
                                titre,
                                date_publication,
                                client,
                                type_appel,
                                date_limite,
                                documents,
                                lien_site
                            ])

                            self.stdout.write(f"Ajouté: {titre[:30]}... | Pub: {date_publication} | Clôture: {date_limite}")

                    except Exception as e:
                        self.stdout.write(f"Erreur sur une ligne: {e}")

                page_number += 1

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Erreur majeure SNIM: {e}"))

        finally:
            wb.save(OUTPUT_FILE)
            self.stdout.write(self.style.SUCCESS(
                f"✅ Fichier Excel SNIM généré: {OUTPUT_FILE}\n"
                f"Contient tous les champs demandés."
            ))

    def scrape_rimtic(self):
        """Fonction de scraping pour RIMTIC avec Selenium"""
        self.stdout.write("\n🔍 Démarrage du scraping RIMTIC...")
        
        OUTPUT_FILE = Path.cwd() / 'scraped_data' / 'rimtic.xlsx'
        OUTPUT_FILE.parent.mkdir(exist_ok=True)

        # Configuration du navigateur
        chrome_options = Options()
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')

        driver = webdriver.Chrome(options=chrome_options)

        try:
            driver.get("https://rimtic.com/")
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            
            try:
                tab_offres = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.ID, "controlled-tab-example-tab-AppelsOffres"))
                )
                driver.execute_script("arguments[0].click();", tab_offres)
                self.stdout.write("Navigation vers les Appels d'Offres réussie")
            except TimeoutException:
                self.stdout.write("Échec de la navigation vers les Appels d'Offres")
                return

            time.sleep(3)

            offer_urls = []
            offer_elements = WebDriverWait(driver, 20).until(
                EC.presence_of_all_elements_located((By.XPATH, "//a[contains(@href, '/fr/appel-offres/')]"))
            )
            
            for elem in offer_elements:
                url = elem.get_attribute('href')
                if url and url not in offer_urls:
                    offer_urls.append(url)
            
            self.stdout.write(f"Nombre d'offres trouvées: {len(offer_urls)}")

            wb = Workbook()
            ws = wb.active
            ws.title = "Appels d'offres RIMTIC"
            ws.append([
                "titre", "date_de_publication", "client", "type_d_appel_d_offre",
                "date_limite", "documents", "lien_site"
            ])

            for url in offer_urls:
                try:
                    driver.get(url)
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.TAG_NAME, 'body'))
                    )
                    time.sleep(1)

                    data = {
                        'titre': "N/A",
                        'date_publication': "N/A",
                        'client': "N/A",
                        'type': "N/A",
                        'date_limite': "N/A",
                        'documents': "N/A",
                        'lien_site': url
                    }

                    try:
                        titre_elem = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, 
                                "//div[contains(@class, 'dangerouslySetInnerHTML') and "
                                "contains(@style, 'text-align: center')]"))
                        )
                        data['titre'] = self.clean_special_characters(titre_elem.text)
                    except (NoSuchElementException, TimeoutException):
                        self.stdout.write(f"⚠️ Titre non trouvé pour {url}")

                    try:
                        client_elem = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH,
                                "//p[contains(@class, 'fontAlmari') and "
                                "contains(@style, 'text-align: center')]"))
                        )
                        data['client'] = self.clean_special_characters(client_elem.text)
                    except (NoSuchElementException, TimeoutException):
                        try:
                            img_elem = WebDriverWait(driver, 5).until(
                                EC.presence_of_element_located((By.XPATH,
                                    "//img[contains(@src, 'storage/images')]"))
                            )
                            data['client'] = self.clean_special_characters(img_elem.get_attribute('alt'))
                        except (NoSuchElementException, TimeoutException):
                            data['client'] = "RIMTIC"

                    try:
                        date_limite_elem = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH,
                                "//p[contains(., 'Limite :')]/span"))
                        )
                        data['date_limite'] = self.clean_special_characters(date_limite_elem.text)
                    except (NoSuchElementException, TimeoutException):
                        pass

                    try:
                        publie_le_elem = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH,
                                "//div[contains(@class, 'div-partager')]//"
                                "div[contains(., 'Publié le :')]"))
                        )
                        data['date_publication'] = self.clean_special_characters(
                            publie_le_elem.text.split(':')[-1].strip())
                    except (NoSuchElementException, TimeoutException):
                        pass

                    try:
                        pdf_elements = WebDriverWait(driver, 5).until(
                            EC.presence_of_all_elements_located((By.XPATH,
                                "//a[contains(@href, '.pdf') and "
                                "contains(@href, 'storage/documents')]"))
                        )
                        pdf_links = [elem.get_attribute('href') for elem in pdf_elements if elem.get_attribute('href')]
                        data['documents'] = ", ".join(pdf_links) if pdf_links else "N/A"
                    except (NoSuchElementException, TimeoutException):
                        pass

                    if data['titre'] != "N/A":
                        titre_lower = data['titre'].lower()
                        if any(word in titre_lower for word in ["consultation", "consultant", "étude"]):
                            data['type'] = "Consultation"
                        elif "avis" in titre_lower:
                            data['type'] = "Avis d'appel d'offres"

                    ws.append([
                        data['titre'],
                        data['date_publication'],
                        data['client'],
                        data['type'],
                        data['date_limite'],
                        data['documents'],
                        data['lien_site']
                    ])

                    self.stdout.write("\n" + "="*50)
                    self.stdout.write(f"Titre: {data['titre']}")
                    self.stdout.write(f"Client: {data['client']}")
                    self.stdout.write(f"Type: {data['type']}")
                    self.stdout.write(f"Publié le: {data['date_publication']}")
                    self.stdout.write(f"Date limite: {data['date_limite']}")
                    self.stdout.write(f"Documents: {data['documents']}")
                    self.stdout.write(f"Lien: {data['lien_site']}")
                    self.stdout.write("="*50)

                except Exception as e:
                    self.stdout.write(f"\n⚠️ Erreur sur la page {url} : {str(e)}")
                    continue

            wb.save(OUTPUT_FILE)
            self.stdout.write(self.style.SUCCESS(
                f"\n✅ Scraping RIMTIC terminé! {len(offer_urls)} appels d'offres enregistrés dans: {OUTPUT_FILE}"))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Erreur globale RIMTIC : {str(e)}"))
        finally:
            driver.quit()



    def scrape_somelec(self):
        """Fonction de scraping pour SOMELEC avec Selenium"""
        self.stdout.write("\n🔍 Démarrage du scraping SOMELEC...")
        
        OUTPUT_FILE = Path.cwd() / 'scraped_data' / 'somelec.xlsx'
        OUTPUT_FILE.parent.mkdir(exist_ok=True)

        chrome_options = Options()
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')

        driver = webdriver.Chrome(options=chrome_options)
        BASE_URL = "https://somelec.mr"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Appels d'offres SOMELEC"
            ws.append([
                "titre", "date_de_publication", "client", "type_d_appel_d_offre",
                "date_limite", "documents", "lien_site"
            ])

            page_number = 0
            while True:
                url = f"{BASE_URL}/?q=ap&page={page_number}"
                self.stdout.write(f"Accès à la page {page_number}...")
                driver.get(url)

                try:
                    WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "article.node-content"))
                    )
                except TimeoutException:
                    self.stdout.write("Timeout lors du chargement des articles.")
                    break

                articles = driver.find_elements(By.CSS_SELECTOR, "article.node-content")
                self.stdout.write(f"Articles trouvés : {len(articles)}")

                if len(articles) == 0:
                    self.stdout.write("Fin de la pagination.")
                    break

                for article in articles:
                    try:
                        titre_element = article.find_element(By.CSS_SELECTOR, "h2.title a")
                        titre = self.clean_special_characters(titre_element.text)
                        detail_url = urljoin(BASE_URL, titre_element.get_attribute("href"))
                        
                        self.stdout.write(f"Accès à la page de détails: {detail_url}")
                        driver.get(detail_url)
                        
                        data = {
                            "titre": titre,
                            "date_de_publication": "N/A",
                            "client": "SOMELEC",
                            "type_d_appel_d_offre": "N/A",
                            "date_limite": "N/A",
                            "documents": [],
                            "lien_site": detail_url
                        }
                        
                        try:
                            WebDriverWait(driver, 30).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "span.date"))
                            )
                            
                            date_selectors = [
                                "span.date",
                                ".submitted .date",
                                ".field-name-post-date",
                                "footer .date",
                                ".node-submitted .date"
                            ]
                            for selector in date_selectors:
                                try:
                                    date_element = driver.find_element(By.CSS_SELECTOR, selector)
                                    date_text = self.clean_special_characters(date_element.text)
                                    if date_text and date_text != "N/A":
                                        data["date_de_publication"] = date_text
                                        break
                                except NoSuchElementException:
                                    continue
                            
                            try:
                                fichiers = driver.find_elements(
                                    By.CSS_SELECTOR, 
                                    ".field-name-field-fichiers-joints a[href$='.pdf']"
                                )
                                data["documents"] = ", ".join(
                                    [urljoin(BASE_URL, f.get_attribute("href")) for f in fichiers]
                                ) if fichiers else "N/A"
                            except NoSuchElementException:
                                data["documents"] = "N/A"
                            
                            if any(mot in data["titre"].lower() for mot in ["consultation", "étude", "avis"]):
                                data["type_d_appel_d_offre"] = "Consultation" if "consultation" in data["titre"].lower() else "Avis d'appel d'offres"
                            
                        except TimeoutException:
                            self.stdout.write("Timeout lors du chargement de la page de détails.")
                        
                        self.stdout.write("\n" + "="*50)
                        self.stdout.write(f"Titre: {data['titre']}")
                        self.stdout.write(f"Date publication: {data['date_de_publication']}")
                        self.stdout.write(f"Client: {data['client']}")
                        self.stdout.write(f"Type: {data['type_d_appel_d_offre']}")
                        self.stdout.write(f"Date limite: {data['date_limite']}")
                        self.stdout.write(f"Documents: {data['documents']}")
                        self.stdout.write(f"Lien: {data['lien_site']}")
                        self.stdout.write("="*50)
                        
                        ws.append([
                            data["titre"],
                            data["date_de_publication"],
                            data["client"],
                            data["type_d_appel_d_offre"],
                            data["date_limite"],
                            data["documents"],
                            data["lien_site"]
                        ])
                        
                        driver.back()
                        WebDriverWait(driver, 30).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "article.node-content"))
                        )
                        
                    except NoSuchElementException as e:
                        self.stdout.write(f"Élément manquant: {str(e)}")
                    except Exception as e:
                        self.stdout.write(f"Erreur lors du traitement de l'article: {str(e)}")
                
                try:
                    next_page = driver.find_element(By.CSS_SELECTOR, "li.pager-next a")
                    page_number += 1
                except NoSuchElementException:
                    self.stdout.write("Pas de page suivante.")
                    break

            wb.save(OUTPUT_FILE)
            self.stdout.write(self.style.SUCCESS(
                f"\nScraping SOMELEC terminé! Données enregistrées dans: {OUTPUT_FILE}"))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Erreur globale SOMELEC: {str(e)}"))
        finally:
            driver.quit()

    def load_all_offers(self, driver):
        """Fonction pour charger toutes les offres en cliquant sur 'Voir plus'"""
        last_count = 0
        same_count = 0
        max_attempts = 30
        
        for attempt in range(1, max_attempts + 1):
            current_offers = driver.find_elements(By.CSS_SELECTOR, "h2.text-sm.mb-2.pt-6.text-green-950")
            current_count = len(current_offers)
            
            self.stdout.write(f"📊 Offres chargées: {current_count} (tentative {attempt}/{max_attempts})")
            
            if current_count == last_count:
                same_count += 1
                if same_count >= 5:
                    self.stdout.write("🔴 Blocage détecté, tentative de déblocage...")
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(3)
                    driver.execute_script("window.scrollTo(0, 0);")
                    time.sleep(2)
                    same_count = 0
            else:
                same_count = 0
                last_count = current_count
            
            try:
                buttons = WebDriverWait(driver, 15).until(
                    EC.presence_of_all_elements_located((By.XPATH, "//button[contains(., 'Voir plus')]"))
                )
                if buttons:
                    last_button = buttons[-1]
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", last_button)
                    time.sleep(1)
                    driver.execute_script("arguments[0].click();", last_button)
                    self.stdout.write(f"✅ Clic sur 'Voir plus' (tentative {attempt})")
                    time.sleep(3)
            except (NoSuchElementException, TimeoutException):
                self.stdout.write("ℹ️ Bouton 'Voir plus' non trouvé - fin du chargement")
                break
            except (ElementClickInterceptedException, StaleElementReferenceException):
                self.stdout.write("⚠️ Clic bloqué, tentative de récupération...")
                time.sleep(3)
                continue
            except Exception as e:
                self.stdout.write(f"❌ Erreur inattendue: {str(e)}")
                time.sleep(3)
                continue


   